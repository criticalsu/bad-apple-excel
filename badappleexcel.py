import os
import cv2 as cv
import openpyxl
import multiprocessing
import time

def rgb_to_hex(array):
    return "{:02x}{:02x}{:02x}".format(array[2], array[1], array[0])

def img_to_excel(files):
    #Read files
    image = cv.imread(f"C:/Users/ACER/Documents/bad apple frame/{files}")
    image_list = image.tolist()
    
    #make worksheet
    wb4 = openpyxl.Workbook()
    ws4 = wb4.active
    #sheet name = frame number
    #ws4.title = f"{str(i).zfill(4)}"

    #convert BGR (wtf opncv) val to hex and append to a list
    image_hex3 = []

    for j in range(360):
        row_hex = []
        for k in range(480):
            row_hex.append("ff" + rgb_to_hex(image[j, k]))
        image_hex3.append(row_hex)
    
    #color each cell according to their respective hex value
    m = 0
    j = 0

    #360 = vertical pixels
    for j in range(360):
        m=0
        for columns in ws4[f"A{j+1}:RL{j+1}"]:
            for cell in columns:
                cell.fill = openpyxl.styles.PatternFill("solid", start_color=image_hex3[j][m])
                m += 1

    #Resize rowheights
    for rows in range(ws4.max_row):
        ws4.row_dimensions[rows+1].height = 11.25

    #Resize colwidths
    for cols in range(480):
        ws4.column_dimensions[f"{openpyxl.utils.get_column_letter(cols+1)}"].width = 2.14

    #save to /frames/ directory (gitignore)
    wb4.save(f"C:/Users/ACER/Documents/GitHub/bad-apple-excel/frames/{files}.xlsx")

start = time.time()

if __name__ == "__main__":
    frames = os.listdir("C:/Users/ACER/Documents/bad apple frame")
    frames360 = frames[:360] #To iterate through all files just delete this line

    p = multiprocessing.Pool()

    p.map(img_to_excel, frames360)

stop = time.time()

print(f"Elapsed: {stop-start}")